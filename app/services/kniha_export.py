"""
services/kniha_export.py — export knihy jízd / cestovního příkazu do Excelu.

Formát kopíruje firemní měsíční šablonu Commarec:
- hlavička: Vozidlo / SPZ / Řidič  |  Firma / IČ / DIČ
- tabulka: Odjezd (Datum, Čas, Odkud) | Příjezd (Datum, Čas, Kam) | Účel
           | Stav tachometru (Začátek, Konec) | Ujeto km | Tankování (litry, Kč)
- řádek Celkem (km, litry, Kč)
Tankování se k jízdě páruje podle data (PHM transakce daného dne).
"""
import io

NAVY = "173767"
SVETLA = "F3F5F7"
CZ_MESICE = ["", "Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
             "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
FIRMA = "Commarec s.r.o."
ICO = "218 36 256"
DIC = "CZ21836256"


def _den(d):
    return d.strftime("%d.%m.%Y") if d else ""


def export_xlsx(vozidlo, rok, mesic, jizdy, tankovani=None, stav_zacatek=None):
    """jizdy = list objektů Jizda (řazené dle data). tankovani = list objektů Tankovani
    daného vozidla a měsíce. Vrací bytes .xlsx ve formátu cestovního příkazu."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{mesic:02d}-{rok}"

    bold = Font(name="Calibri", bold=True)
    navy_bold = Font(name="Calibri", bold=True, color=NAVY)
    hl_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hl_fill = PatternFill("solid", fgColor=NAVY)
    zebra = PatternFill("solid", fgColor=SVETLA)
    tenky = Side(style="thin", color="C9D2DE")
    ramecek = Border(left=tenky, right=tenky, top=tenky, bottom=tenky)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    levy = Alignment(horizontal="left", vertical="center")

    # ── titulek ──
    ws["A1"] = "Kniha jízd / cestovní příkaz"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws["A2"] = f"{CZ_MESICE[mesic]} {rok}"
    ws["A2"].font = Font(bold=True, size=11)

    # ── hlavička: vozidlo (vlevo) + firma (vpravo) ──
    hlav = [
        ("Vozidlo", vozidlo.model or "", "Firma", FIRMA),
        ("SPZ", vozidlo.spz or "", "IČ", ICO),
        ("Řidič", vozidlo.ridic or "", "DIČ", DIC),
    ]
    r = 4
    for popis_l, hod_l, popis_p, hod_p in hlav:
        ws.cell(row=r, column=1, value=popis_l).font = bold
        ws.cell(row=r, column=3, value=hod_l)
        ws.cell(row=r, column=7, value=popis_p).font = bold
        ws.cell(row=r, column=8, value=hod_p)
        r += 1

    # ── záhlaví tabulky (dvouřádkové) ──
    hr1 = r + 1
    skupiny = [(1, "Odjezd"), (4, "Příjezd"), (7, "Účel cesty"),
               (8, "Stav tachometru"), (10, "Ujeto km"), (11, "Tankování")]
    for col, nazev in skupiny:
        c = ws.cell(row=hr1, column=col, value=nazev)
        c.font = hl_font
        c.fill = hl_fill
        c.alignment = center
    ws.merge_cells(start_row=hr1, start_column=1, end_row=hr1, end_column=3)   # Odjezd
    ws.merge_cells(start_row=hr1, start_column=4, end_row=hr1, end_column=6)   # Příjezd
    ws.merge_cells(start_row=hr1, start_column=7, end_row=hr1 + 1, end_column=7)  # Účel
    ws.merge_cells(start_row=hr1, start_column=8, end_row=hr1, end_column=9)   # Stav tachometru
    ws.merge_cells(start_row=hr1, start_column=10, end_row=hr1 + 1, end_column=10)  # Ujeto km
    ws.merge_cells(start_row=hr1, start_column=11, end_row=hr1, end_column=12)  # Tankování

    hr2 = hr1 + 1
    podzahlavi = {1: "Datum", 2: "Čas", 3: "Odkud", 4: "Datum", 5: "Čas", 6: "Kam",
                  8: "Začátek", 9: "Konec", 11: "litry", 12: "Kč"}
    for col, nazev in podzahlavi.items():
        c = ws.cell(row=hr2, column=col, value=nazev)
        c.font = hl_font
        c.fill = hl_fill
        c.alignment = center
    for col in range(1, 13):  # rámeček pod celé záhlaví
        ws.cell(row=hr1, column=col).border = ramecek
        ws.cell(row=hr2, column=col).border = ramecek

    # ── párování tankování k datu (PHM) ──
    tank_dle_data = {}
    for t in (tankovani or []):
        if t.datum and t.kategorie != "ostatni":
            tank_dle_data.setdefault(t.datum, []).append(t)

    # ── řádky jízd ──
    stav = stav_zacatek if stav_zacatek is not None else (vozidlo.tachometr_pocatek or 0)
    soucet_km = soucet_l = soucet_kc = 0.0
    rr = hr2 + 1
    for i, j in enumerate(jizdy):
        km = float(j.km or 0)
        zac = stav
        kon = stav + km
        stav = kon
        soucet_km += km
        # tankování daného dne (vyber jedno nepoužité)
        litry = castka = None
        seznam = tank_dle_data.get(j.datum) or []
        for t in seznam:
            if not getattr(t, "_pouzito", False):
                t._pouzito = True
                litry, castka = t.litry, t.castka
                if litry:
                    soucet_l += float(litry)
                if castka:
                    soucet_kc += float(castka)
                break
        radek = [_den(j.datum), "", j.odkud or "", _den(j.datum), "", j.kam or "",
                 "soukromě" if j.soukroma else (j.ucel or "služebně"),
                 round(zac), round(kon), round(km), litry or "", castka or ""]
        for c, val in enumerate(radek, start=1):
            bunka = ws.cell(row=rr, column=c, value=val)
            bunka.border = ramecek
            bunka.alignment = levy if c in (3, 6, 7) else center
            if i % 2 == 1:
                bunka.fill = zebra
        rr += 1

    # ── Celkem ──
    ws.cell(row=rr, column=9, value="Celkem").font = bold
    ws.cell(row=rr, column=10, value=round(soucet_km)).font = bold
    ws.cell(row=rr, column=11, value=round(soucet_l, 2) if soucet_l else "").font = bold
    ws.cell(row=rr, column=12, value=round(soucet_kc, 2) if soucet_kc else "").font = bold
    for col in range(1, 13):
        ws.cell(row=rr, column=col).border = ramecek

    # ── šířky sloupců ──
    sirky = {"A": 11, "B": 7, "C": 16, "D": 11, "E": 7, "F": 16, "G": 14,
             "H": 10, "I": 10, "J": 9, "K": 8, "L": 10}
    for col, w in sirky.items():
        ws.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
